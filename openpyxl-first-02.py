"""
Please note, the excessive amount of comments are due to
how cofusing the api is and the specific cells in the output excel.
"""
from openpyxl import load_workbook
import requests
from urllib.parse import quote_plus
from urllib.request import urlopen
from urllib.error import HTTPError
from bs4 import BeautifulSoup
import pandas as pd
import random
import time


LIST_PRINT_TOP10 = []
LIST_ONLINE_TOP10 = []

BITLY_API = 'e8c05d9133d039c84ab05395e7871e0c0e3b2432'

SUBJECT_LIST = ['GT-R', 'NV 350 Impendulo', 'Qashqai', 'NV 200 Combi',
                'Tiida', 'Leaf', 'NAAMSA', 'K-line', 'Juke', 'X-Trail',
                'Navara', 'Micra', 'NP 200', 'Patrol', 'Sentra',
                'Nissan Festival of Motoring',
                'Nissan drives for further growth',
                'ICC Sponsorship', 'NP 300', 'Nissan Trailseeker',
                'Nissan Corporate',
                'Nissan Taxi Call Centre oppurtunity',
                'Nissn long distance driving tips',
                'Nissan Holiday Checklist', 'Pathfinder', 'NP 300 hardbody',
                'NV 300', 'Workhorse', 'Almera', '350z', 'Hardbody',
                'Simola Hillclimb', '370z', 'Murano',
                'Nissan Kicks', 'Nissan Primera', 'Bladeglider',
                'Nissan ProPilot', 'Festival of Motoring', 'NV 350 Panel Van',
                'Patrol Wagon', 'Livina', 'Nissan Sponsorship/sport',
                'Nissan Motorsport','Nissan Spokespeople',
                'General Industry related']


def loadPrime():
    primeList = ('X:/DDI Media Monitoring/Data for Analysis/Nissan/2018/' \
                 'Nissan Weekly/Prime Reading Current as of 02.2018.xlsx')
    
    df = pd.read_excel(primeList)
    
    printList = list(df['PRINT '])
    printList = list(filter(lambda x: x==x, printList))
    
    onlineList = list(df['ONLINE '])
    onlineList = list(filter(lambda x: x==x, onlineList))
    
    tvList = list(df['TV'])
    tvList = list(filter(lambda x: x==x, tvList))

    return printList, onlineList, tvList


PRINT_LIST, ONLINE_LIST, TV_LIST = loadPrime()


def findSubject(text, description):
    productFound = ''
    for product in SUBJECT_LIST:
        if (text.lower().find(product.lower()) != -1) or \
           (description.lower().find(product.lower()) != -1):
            productFound = product
    
    return productFound


def findPrimePrint(media):
    primeString = ''
    for prime in PRINT_LIST:
        if media.lower().find(prime.lower()) != -1:
            primeString = ' [PRIME]'

    return primeString


def findPrimeOnline(media):
    primeString = ''
    for prime in ONLINE_LIST:
        if media.lower().find(prime.lower()) != -1:
            primeString = ' [PRIME]'

    return primeString


def findPrimeBroadcast(media):
    primeString = ''
    for prime in TV_LIST:
        if media.lower().find(prime.lower()) != -1:
            primeString = ' [PRIME]'

    return primeString
    
    
def getExtract(url, description):
    text = description

    try:
        page = urlopen(url)
        soup = BeautifulSoup(page, 'html.parser')
        paragraphs = soup.find_all('p')

        for paragraph in paragraphs:
            text += paragraph.get_text() + ' '

        while len(text) > 700:
            index = text.rfind('.')
            text = text[:index]

        text += '.'

    except HTTPError:
        text = description
        
    return text


def printRow(row, coverage):
    #Converts the url
    url = quote_plus(row[4].value)
    
    #Bitly GET call address
    api_call = 'https://api-ssl.bitly.com/v3/shorten?access_token={}'\
           '&longUrl={}'.format(BITLY_API, url)
    
    #Using requests for the GET call
    url = requests.get(api_call).json()['data']['url'] 

    #These variables are added for readability

    title = row[0].value
    media = row[1].value + findPrimePrint(row[1].value)
    date = row[2].value
    tone = row[10].value
    reach = row[6].value
    ave = row[9].value
    link = url
    extract = row[5].value.replace('_x000D_\n',
                                   ' ').replace('\ufeff','')

    count = 0
    while len(extract)-1 > 500:
        index = extract.rfind('.')
        extract = extract[:index]
        count += 1

    extract = extract + '.'

    if coverage == 'Product':
        nameplate = row[18].value
    else:
        nameplate = findSubject(extract, row[5].value)

    nameplate = nameplate.replace(', Nissan', '')

    #Might use the code below in the future
    #commaIndex = row[16].value.find(',')
    #subjects = row[16].value[:4]

    return [nameplate, title, media, date, tone, reach, ave, link, extract]


def onlineRow(row, coverage):
    """
    parses row of sheet
    """
    #Converts the url
    url = quote_plus(row[4].value)

    #Bitly GET call address
    api_call = 'https://api-ssl.bitly.com/v3/shorten?access_token={}'\
           '&longUrl={}'.format(BITLY_API, url)

    #Using requests for the GET call
    url = requests.get(api_call).json()['data']['url'] 

    #These variables are added for readability
    title = row[2].value
    

    if coverage == 'SSA Online':
        media = row[12].value
        index = media.find('//')
        media = media[index+2:]
        media = media.replace('www.','')
        media = media + findPrimeOnline(media) 
        tone = row[10].value
        reach = row[18].value
        ave = row[19].value
        #Get Longer extract of description
        extract = getExtract(row[4].value, row[3].value)
        if extract == '':
            row[3].value
    elif coverage == 'Social':
        media = row[11].value
        tone = row[11].value
        reach = row[17].value
        ave = row[19].value
        extract = row[2].value
        index = title.find('http')
        title = extract[:index-1]
    else: #  Online
        media = row[13].value
        index = media.find('//')
        media = media[index+2:]
        media = media.replace('www.','')
        media = media + findPrimeOnline(media) 
        tone = row[11].value
        reach = row[19].value
        ave = row[20].value
        extract = getExtract(row[4].value, row[3].value)
        if len(extract) < 100:
            row[3].value
    
    date = row[5].value[:10]

    nameplate = findSubject(extract, row[2].value) 

    link = url

    return [nameplate, title, media, date, tone.capitalize(), reach, ave, link, extract]


def copyTop5(wb, ws1, ws2, coverage, media):
    #Colours of the highlighted cells
    colours = {'Product' : 'FFFFFF00',  'Corporate' : 'FF92D050',
               'SSA Online': 'FFFFC000', 'Social': 'FF33CAFF'}
    colour = colours[coverage]
    
    #creates list of rows from ws1
    rows = list(ws1.rows)

    #creates list of thrid column in ws2
    column = list(ws2.columns)[2]

    #Increases by 11 each loop because of the excel sheets setup 
    clip_count = 0

    #Count the number of clips added
    count = 0 

    #iterate first cell in each row until you find a yellow cell 
    for row in rows:
        #If the cell is a certan colour
        if row[0].fill.fgColor.rgb == colour:

            #Max 5 clips
            if (count==3) and (coverage=='Social'):
                break
            elif count==5:
                break

            if media=='Print':
                rowValues = printRow(row, coverage)
            else:
                rowValues = onlineRow(row, coverage)
   
            column[1+clip_count].value = rowValues[0] #nameplate
            column[2+clip_count].value = rowValues[1] #title
            column[3+clip_count].value = rowValues[2] #media
            column[4+clip_count].value = rowValues[3] #date
            column[5+clip_count].value = rowValues[4] #tone
            column[6+clip_count].value = rowValues[5] #reach
            column[7+clip_count].value = rowValues[6] #ave
            column[8+clip_count].value = rowValues[7] #link
            column[9+clip_count].value = rowValues[8] #extract 
            
            clip_count += 11
            count+=1

            if media == 'Print':
                LIST_PRINT_TOP10.append([rowValues[2], rowValues[1]])
            else:
                LIST_ONLINE_TOP10.append([rowValues[2], rowValues[1]])
            
    wb.save('done.xlsx')


def copyBroadcast(wb, ws1, ws2, coverage, media):
    #Colour of the highlighted cells
    colour = 'FFFF0000'
    
    #creates list of rows from ws1
    rows = list(ws1.rows)

    #creates list of thrid column in ws2
    column = list(ws2.columns)[2]

    #Increases by 11 each loop because of the excel sheets setup 
    clip_count = 0

    #Count the number of clips added
    count = 0 

    #iterate first cell in each row until you find a yellow cell 
    for row in rows:
        #If the cell is yellow
        #if row[0].fill.fgColor.rgb != '0000000':
        if row[0].fill.fgColor.rgb == colour:

            #Max 5 clips
            if count==2:
                break

            #These variables are added for readability
            title = row[1].value
            
            media = str(row[3].value)
            date = row[4].value
            tone = row[18].value
            reach = row[13].value
            ave = row[12].value
            extract = row[10].value
            nameplate = row[9].value

            column[1+clip_count].value = nameplate                  #nameplate
            column[2+clip_count].value = title                      #title
            column[3+clip_count].value = media + \
                                         findPrimeBroadcast(media)  #media
            column[4+clip_count].value = date                       #date
            column[5+clip_count].value = tone                       #tone
            column[6+clip_count].value = reach                      #reach
            column[7+clip_count].value = ave                        #ave
            column[8+clip_count].value = extract                    #extract 
            
            clip_count += 10
            count+=1

    wb.save('done.xlsx')

def checkIfPrime(media):
    """Check if it's prime.
    """
    if media.find == 'PRIME':
        return True
    else:
        return False
    

def top10Stories(wb, ws):
    """Generates Top10 from print and online clips
    """
    ws = ws['J11':'L20']

    #Randomly selects 5 print clips for print and online
    selection1 = random.sample(range(0, 10), 5)
    selection2 = random.sample(range(0, 10), 5)

    for index in range(5):
        if checkIfPrime(LIST_PRINT_TOP10[selection1[index]][0]):
            ws[index][2].value = 'PRIME'

        media = LIST_PRINT_TOP10[selection1[index]][0].replace(' [PRIME]', '')
        
        ws[index][0].value = media                                    #media
        ws[index][1].value = LIST_PRINT_TOP10[selection1[index]][1]   #title

    for index in range(5,10):
        if checkIfPrime(LIST_ONLINE_TOP10[selection2[index-5]][0]):
            ws[index][2].value = 'PRIME'

        media = LIST_ONLINE_TOP10[selection2[index-5]][0].replace(' [PRIME]',
                                                                  '')
        
        ws[index][0].value = media                                     #media
        ws[index][1].value = LIST_ONLINE_TOP10[selection2[index-5]][1] #title

    wb.save('done.xlsx')
    

def main():
    start_time = time.time()
    #Load excel document
    wb = load_workbook('AVEsmall.xlsx', data_only=True)

    #Load Print sheets
    wsPrint = wb['Print Top AVE']
    wsProductPrint = wb['PRINT PRODUCT']
    wsCorporatePrint = wb['PRINT CORPORATE']

    #Copy Print Sheets to Coverage Highlights
    copyTop5(wb, wsPrint, wsProductPrint, 'Product', 'Print')
    copyTop5(wb, wsPrint, wsCorporatePrint, 'Corporate', 'Print')

    #Load Online sheets
    wsOnline = wb['Online Top AVE']
    wsProductOnline = wb['PRODUCT ONLINE']
    wsCorporateOnline = wb['CORPORATE ONLINE']

    #Copy Online Sheets to Coverage Highlights
    copyTop5(wb, wsOnline, wsProductOnline, 'Product', 'Online')
    copyTop5(wb, wsOnline, wsCorporateOnline, 'Corporate', 'Online')

    #Load Broadcast sheets
    wsBroadcastList = wb['Broadcast']
    wsBroadcast = wb['BROADCAST TABLE DETAIL']

    #Copy Broadcast Sheets to Coverage Highlights
    copyBroadcast(wb, wsBroadcastList, wsBroadcast, 'Broadcast', 'Broadcast')


    #Load SSA sheets
    wsSSA = wb['SSA Online']
    wsSSAOnline = wb['SSA Online Details']
    
    wsSocial = wb['SSA Social Media']
    wsSocialOnline = wb['SSA SM Details']

    #Copy SSA Sheets to Coverage Highlights
    copyTop5(wb, wsSSA, wsSSAOnline, 'SSA Online', 'Online')
    copyTop5(wb, wsSocial, wsSocialOnline, 'Social', 'Social')


    #Copy Top 10 Stories
    wsTop10 = wb['Overall Summary Slide 2']
    top10Stories(wb, wsTop10)

    print("--- Took %s seconds ---" % (time.time() - start_time))
    

if __name__ == "__main__":
    main()
