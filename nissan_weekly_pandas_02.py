"""
.. module:: nissan_weekly
   :platform: Windows
   :synopsis: A useful module indeed.

.. moduleauthor:: Xolani Mazibuko <xolani@ddi.media>

"""
import pandas as pd
import numpy as np
import requests


from urllib.parse import quote_plus
from urllib.error import HTTPError
from urllib.error import URLError
from urllib.request import urlopen

from bs4 import BeautifulSoup
from openpyxl import load_workbook

import datetime
import time


subjectList = ['GT-R', 'NV 350 Impendulo', 'Qashqai', 'NV 200 Combi',
               'Tiida', 'Leaf', 'NAAMSA', 'K-line', 'Juke', 'X-Trail',
               'Navara', 'Micra', 'NP 200', 'Patrol', 'Sentra',
               'Nissan Festival of Motoring',
               'Nissan drives for further growth',
               'ICC Sponsorship', 'NP 300', 'Nissan Trailseeker',
               'Nissan Corporate',
               'Nissan Taxi Call Centre oppurtunity',
               'Nissn long distance driving tips',
               'Nissan Holiday Checklist', 'Pathfinder', 'NP 300 hardbody',
               'NV 300', 'Workhorse', 'Almera', '350z', 'Hardbody',
               'Simola Hillclimb', '370z', 'Murano', 'Nissan',
               'Nissan Kicks', 'Nissan Primera', 'Bladeglider',
               'Nissan ProPilot', 'Festival of Motoring', 'NV 350 Panel Van',
               'Patrol Wagon', 'Livina', 'Nissan Sponsorship/sport',
               'Nissan Motorsport','Nissan Spokespeople',
               'General Industry related']

productList = ['NP 300', 'GT-R', 'Qashqai', 'NV 200 Combi', 'Tiida',
               'NV 300',  'Leaf', 'K-line', 'Juke', 'X-trail', 'NP 300',
               '350z', 'Navara', 'Micra', 'NP 200', 'Patrol', 'Sentra' ,
               '370z', 'Murano', 'Pathfinder', 'NP 300 hardbody', 'NV 300',
               'Workhorse', 'Almera', 'NV 350 Panel Van', 'Livina',
               'Patrol Wagon', 'Nissan Primera', 'Nissan ProPilot',
               'Bladeglider']

corporateList = ['NAAMSA', 'Nissan Corporate', 'Nissan', 'Nissan Motorsport',
                 'Nissan Spokespeople', 'General Industry related',
                 'Nissan Sponsorship/sport', 'Festival of Motoring',
                 'Nissn long distance driving tips',
                 'Nissan Holiday Checklist', 'Nissan Taxi Call Centre oppurtunity',
                 'ICC Sponsorship', 'Nissan Festival of Motoring',
                 'Nissan drives for further growth', 'Nissan Trailseeker']

DELETE_LIST = []


def inputfile(excelFile):
    """

    :param excelFile:
    :return:
    """
    df = pd.read_excel(excelFile)
    return df


def inputfilePrint(excelFile):
    """Takes in an excel file name and returns a DataFrame without the
        first two rows.

    Args:
       excelFile (str):  The excel files name.

    Returns:
       class DataFrame. 
    """
    df = pd.read_excel(excelFile, skiprows=2)
    return df


def outputfile(df1, df2, df3, excelFileOut='final.xlsx'):
    """
    Takes in an DataFrame, the columns and new excel file name.

    Outputs the new excel file with the new excl file name.
    """
    sheetNamePrint = 'Combined'

    
    

    writer = df.to_excel(excelFileOut, index=False)
    df1.to_excel(writer, 'Print')
    df2.to_excel(writer, 'Online')
    df3.to_excel(writer, 'Broadcast')
    writer.save()
         
    print(excelFileOut, 'is done!')


def get_outputfile_name():
    """
    """
    date = ('AVE Report ' +
            (datetime.date.today()-datetime.timedelta(days=7)).strftime(
                                                                    "%d.%m.%Y")
            + ' - ' +
            datetime.date.today().strftime("%d.%m.%Y") +
            '.xlsx')

    return date


def populate_delete_list():
    """Populates the DELETE_LIST
    """
    

def check_url(url):
    """Checks if the given url is in the auto delete list.
    """
    
    if DELETE_LIST.find(url):
        return True
    else:
        return False

def getExtract(url):
    """Returns the paragraphs from a webpage
    """
    text = ''

    try:
        page = urlopen(url)
        soup = BeautifulSoup(page, 'lxml')
        paragraphs = soup.find_all('p')

        for paragraph in paragraphs:
            text += paragraph.get_text(strip=True).replace(u'\xa0', u' ') + ' '

        while len(text) > 500:
            index = text.rfind('.')
            text = text[:index]

        text += '.'

    except (HTTPError, URLError) as error:
        text = ''
        
    return text


def find_category(description, url):
    """Determines whether the category is Product or Corporate from description
        and webpage paragraphs.
    """

    for product in productList:
        if product.lower() in description.lower():
            return True
        else:
            text = getExtract(url)
            if text != '':
                if product.lower() in text.lower():
                    return True
                else:
                    return False
            else:
                return False
    

def find_domain(url):
    """Determines domain_reach from excelfile with domain numbers
    """

    for index, row in df.iterrows():
        if url == df.urlList[index]:
            return df.reach[index]
        else:
            return 1000


def online_fix(df):
    """
    """
    start_time = time.time()

    df = df.rename(columns = {'favorite' : 'category',
                              'tags' : 'AVE',
                              })

    #  Capatilize first letter of Sentiment
    df.tone = df.tone.str.capitalize()

    #  Converts the values to datetime values
    df.published_at = pd.to_datetime(df.published_at)
    #  Gets rid of the time without time
    df.published_at = df.published_at.dt.date

    
    
    for index, row in df.iterrows():
##        if check_url:
##            df.drop(index, inplace=True)
##            df = df.reset_index()

##        if find_category(row.description, row.url):
##            df.loc[[index], 'category'] = 'Product'
##        else:
##            df.loc[[index], 'category'] = 'Corporate'

        if (row.domain_reach == 0) or (np.isnan(row.domain_reach)):
            df.loc[[index], 'domain_reach'] = 1000 ##  find_domain(row.url)

    df.AVE = df.domain_reach*0.21

    df = df.drop(['id', 'alert_id', 'parent_id', 'children', 'parent_url',
                  'direct_reach', 'cumulative_reach', 'alert_name'],
                 axis=1)

    print("--- Took %s seconds ---" % (time.time() - start_time))

    return df
        
        
    

    
    
    


def print_fix(df):
    """Changes values for each row.
    """
    
    df['Category'] = ''


    #  Determines whether the category is Product or Corporate
    category = []
    for index, row in df.iterrows():
        category = []
        subList = df.Subjects[index]
        for sub in subList.split(', '):
            #  Dirty code!!!
            try:
                if productList.index(sub) != -1:
                    cat = 'Product'
                    break
            except ValueError:
                cat = 'Corporate'
        category.append(cat)
        df.loc[[index], 'Category'] = category[0]

##        if sub in productList:
##            df.loc[[index], 'Category'] = 'Product'
            
    
    #  Change Sentiment from number to String
    for index, row in df.iterrows():
        if df.Sentiment[index] == 1:
            df.loc[[index], 'Sentiment'] = "Positive"
        elif df.Sentiment[index] == 0:
            df.loc[[index], 'Sentiment'] = "Neutral"
        elif df.Sentiment[index] == -1:
            df.loc[[index], 'Sentiment'] = "Negative"

    #  Converts the values to datetime values
    df['Referred Date'] = pd.to_datetime(df['Referred Date'])
    #  Gets rid of the time
    df['Referred Date'] = df['Referred Date'].dt.date

    df = df.drop(['Scanned Date', 'Reach', 'Language', 'Curation Date',
                  'Client'],
                 axis=1)
                              
    return df


def broadcast_fix_old(df):
    """
    """
    
    df.Date  = pd.to_datetime(df.Date)
    df.Date = df.Date.dt.date

    nissan = df['Client'] == 'Nissan'
    week = df['Date'] >= (datetime.date.today() - datetime.timedelta(days=7))

    df = df[nissan & week].reset_index()
    df['Total AVE'] = df['Total AVE'].astype(str).str[1:]
    df['Total AVE'] = df['Total AVE'].str.replace(',','')
    df['Total AVE'] = pd.to_numeric(df['Total AVE'])

    df = df.drop(['Tagger', 'Client', 'FTP Export', 'Timestamp'], axis=1)

    return df



    

def copysnapshotData(excelFile):
    wb = load_workbook(excelFile)

    
    


printFile = 'print.xlsx'
broadcastFile = 'broadcast.xlsx'
onlineFile = 'onlineandsocial.xlsx'



outputfile = 'AVE Report 05.04.2018 - 11.04.2018.xlsx' # get_outputfile_name()

dfPrint = inputfilePrint(printFile)
dfOnline = inputfile(onlineFile)
dfBroadcast = inputfile(broadcastFile)

dfPrint = print_fix(df)
df = online_fix(df)
dfBroadcast = broadcast_fix_old(df)

outputfile(dfPrint, dfOnline, dfBroadcast, outputfile)

wb = load_workbook(outputfile, data_only=True)
ws = wb['Summary']
ws['B9'].value = (datetime.date.today()-datetime.timedelta(days=7)).strftime(
                                                                    "%d.%m.%Y")
                    + ' - ' + datetime.date.today().strftime("%d.%m.%Y"))

TV = dfBroadcast['Platform'] == 'TV'
Radio = dfBroadcast['Platform'] == 'Radio'

ws['D12'].value = dfPrint.Ave.sum()
ws['D13'].value = dfOnline.Ave.sum()


ws['D14'].value = dfBroadcast.Ave.sum()
ws['D15'].value = dfPrint.Ave.sum()

ws['E12'].value = dfPrint.Ave.sum()
ws['E13'].value = dfPrint.Ave.sum()
ws['E14'].value = dfPrint.Ave.sum()
ws['E15'].value = dfPrint.Ave.sum()

pivotBroadcast = pd.pivot_table(df, index = ['Station'],
                                values = ['Total AVE'],
                                aggfunc={'Total AVE':
                                         [np.sum, 'count']}).sort_values(by=['sum'],
                                                                         ascending=False)



wb.save(outputfile)
