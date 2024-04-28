"""
.. module:: nissan_weekly
   :platform: Windows
   :synopsis: A useful module indeed.

.. moduleauthor:: Xolani Mazibuko <xolani@ddi.media>

"""

import datetime
import time
import types

# <----------Patch-------------->
from itertools import product
from urllib.error import HTTPError, URLError
from urllib.parse import quote_plus
from urllib.request import urlopen

import numpy as np
import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import drawing, load_workbook, worksheet
from openpyxl.utils import range_boundaries


def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(
        self,
        range_string=None,
        start_row=None,
        start_column=None,
        end_row=None,
        end_column=None,
    ):
        """Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = "%s%s:%s%s" % (
                get_column_letter(start_column),
                start_row,
                get_column_letter(end_column),
                end_row,
            )
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A1:E1)")
        else:
            range_string = range_string.replace("$", "")

        if range_string not in self.merged_cells:
            self.merged_cells.add(range_string)

        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+1)
        # cols = range(min_col, max_col+1)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        # for c in islice(cells, 1, None):
        # if c in self._cells:
        # del self._cells[c]

    # Apply monkey patch
    worksheet.Worksheet.merge_cells = merge_cells


patch_worksheet()
# <----------Patch------------->


subjectList = [
    "GT-R",
    "NV 350 Impendulo",
    "Qashqai",
    "NV 200 Combi",
    "Tiida",
    "Leaf",
    "NAAMSA",
    "K-line",
    "Juke",
    "X-Trail",
    "Navara",
    "Micra",
    "NP 200",
    "Patrol",
    "Sentra",
    "Nissan Festival of Motoring",
    "Nissan drives for further growth",
    "ICC Sponsorship",
    "NP 300",
    "Nissan Trailseeker",
    "Nissan Corporate",
    "Nissan Taxi Call Centre oppurtunity",
    "Nissn long distance driving tips",
    "Nissan Holiday Checklist",
    "Pathfinder",
    "NP 300 hardbody",
    "NV 300",
    "Workhorse",
    "Almera",
    "350z",
    "Hardbody",
    "Simola Hillclimb",
    "370z",
    "Murano",
    "Nissan",
    "Nissan Kicks",
    "Nissan Primera",
    "Bladeglider",
    "Nissan ProPilot",
    "Festival of Motoring",
    "NV 350 Panel Van",
    "Patrol Wagon",
    "Livina",
    "Nissan Sponsorship/sport",
    "Nissan Motorsport",
    "Nissan Spokespeople",
    "General Industry related",
]

productList = [
    "NP 300",
    "GT-R",
    "Qashqai",
    "NV 200 Combi",
    "Tiida",
    "NV 300",
    "Leaf",
    "K-line",
    "Juke",
    "X-trail",
    "NP 300",
    "350z",
    "Navara",
    "Micra",
    "NP 200",
    "Patrol",
    "Sentra",
    "370z",
    "Murano",
    "Pathfinder",
    "NP 300 hardbody",
    "NV 300",
    "Workhorse",
    "Almera",
    "NV 350 Panel Van",
    "Livina",
    "Patrol Wagon",
    "Nissan Primera",
    "Nissan ProPilot",
    "Bladeglider",
]

corporateList = [
    "NAAMSA",
    "Nissan Corporate",
    "Nissan",
    "Nissan Motorsport",
    "Nissan Spokespeople",
    "General Industry related",
    "Nissan Sponsorship/sport",
    "Festival of Motoring",
    "Nissn long distance driving tips",
    "Nissan Holiday Checklist",
    "Nissan Taxi Call Centre oppurtunity",
    "ICC Sponsorship",
    "Nissan Festival of Motoring",
    "Nissan drives for further growth",
    "Nissan Trailseeker",
]


def populate_delete_list(deletelist="delete_list.csv"):
    """Populates the DELETE_LIST

    Args:
        deletelist (string): The name of the delete file

    Returns:
       Returns List
    """
    deleteList = []
    for url in list(pd.read_csv(deletelist)["Delete List"]):
        deleteList.append(
            url.replace("http://", "").replace("https://", "").replace("www.", "")
        )

    return deleteList


def url_reach(url_reach_list="data-files/URL Database Master.xlsx"):
    """Extracts domain reach from excel file

    Args:
        url_reach_list (string): The name of the url reach file

    Returns:
           url_reach (dict)
    """
    deleteList = []

    df = pd.read_excel(url_reach_list)

    df = df.dropna()

    for index, row in df.iterrows():
        url = row.url.replace("https://", "").replace("http://", "").replace("www.", "")
        url = url[: url.find("/")]
        df.loc[[index], "url"] = url

    df = df.drop_duplicates(subset="url", keep="last")
    df = df.drop(columns="AVE")
    df = df.set_index("url")

    url_reach = df.to_dict()["domain_reach"]

    return url_reach


def inputfile(excelFile):
    """Takes in an excel file name and returns a DataFrame.

    Args:
       excelFile (str):  The excel files name

    Returns:
       df (pandas.DataFrame)
    """
    df = pd.read_excel(excelFile)
    return df


def inputfilePrint(excelFile):
    """Takes in an excel file name and returns a DataFrame without the
        first two rows.

    Args:
       excelFile (str):  The input excel files name

    Returns:
       df (pandas.DataFrame)
    """
    df = pd.read_excel(excelFile, skiprows=2)
    return df


def outputfile(df1, df2, df3, wb, excelFileOut):
    """Creates an output excel file using the 3 dataframes.

    Args:
       df1 (pandas.DataFrame): DataFrame to combine
       df2 (pandas.DataFrame): DataFrame to combine
       df3 (pandas.DataFrame): DataFrame to combine
       wb (openpyxl.workbook): Excel workbook to combine
       excelFile (str):  The output excel files name.
    """

    writer = pd.ExcelWriter(excelFileOut, engine="openpyxl")
    writer.book = wb
    writer.sheets = {wb["Summary"].title: wb["Summary"]}

    df1.to_excel(writer, "Print", index=False)
    df2.to_excel(writer, "Online", index=False)
    df3.to_excel(writer, "Broadcast", index=False)
    writer.save()

    print(excelFileOut, "is done!")


def outputfileSSA(df1, df2, excelFileOut):
    """Creates an output excel file using the 3 dataframes.

    Args:
       df1 (pandas.DataFrame): DataFrame to combine
       df2 (pandas.DataFrame): DataFrame to combine
       excelFile (str):  The output excel files name.
    """

    writer = pd.ExcelWriter(excelFileOut)
    df1.to_excel(writer, "Online SSA", index=False)
    df2.to_excel(writer, "Social SSA", index=False)
    writer.save()

    print(excelFileOut, "is done!")


def get_outputfile_name():
    """Generates a string using today's date and a date from 7 days ago

    Returns:
       date (string)
    """
    date = (
        "AVE Report "
        + (datetime.date.today() - datetime.timedelta(days=7)).strftime("%d.%m.%Y")
        + " - "
        + (datetime.date.today() - datetime.timedelta(days=1)).strftime("%d.%m.%Y")
        + ".xlsx"
    )

    return date


def get_outputfile_nameSSA():
    """Generates a string using today's date and a date from 7 days ago

    Returns:
       date (string)
    """
    date = (
        "SSA digital data for weekly report "
        + (datetime.date.today() - datetime.timedelta(days=7)).strftime("%d.%m.%Y")
        + " - "
        + (datetime.date.today() - datetime.timedelta(days=1)).strftime("%d.%m.%Y")
        + ".xlsx"
    )

    return date


def check_url(url, deleteList):
    """Checks if the given url is in the auto delete list.

    Args:
        url (string): URL to compare with delete list
        deleteList (:obj: 'list' of :obj: 'str'): List of URL

    Returns:
       bool: True if url is in the list
    """
    url = (
        url.replace("http://", "")
        .replace("https://", "")
        .replace("www.", "")
        .replace("www.", "")
    )
    if url in deleteList:
        return True
    else:
        return False


def getExtract(url):
    """Returns the paragraphs from a webpage

    Args:
        url (string): URL to extract paragraphs from

    Returns:
       text (str): Text from webpage paragraphs
    """
    text = ""

    try:
        page = urlopen(url)
        soup = BeautifulSoup(page, "lxml")
        paragraphs = soup.find_all("p")

        for paragraph in paragraphs:
            text += paragraph.get_text(strip=True).replace("\xa0", " ") + " "

        while len(text) > 500:
            index = text.rfind(".")
            text = text[:index]

        text += "."

    except (HTTPError, URLError) as error:
        text = ""

    return text


def find_category(title, description, url):
    """Determines whether the category is Product or Corporate from description
        and webpage paragraphs.

    Args:
        description (string): Text from Online Data Frame
        url (string): URL to extract paragraphs from

    Returns:
       bool: True if category is found
    """
    # Definately a cleaner and better way to do this
    for product in productList:
        if product.lower() in title.lower():
            return True

    for product in productList:
        if product.lower() in str(description).lower():
            return True

    text = getExtract(url)

    for product in productList:
        if text != "":
            if product.lower() in text.lower():
                return True
            else:
                return False
        else:
            return False


def find_domain(url, url_list):
    """Determines domain_reach from excelfile with domain numbers

    Args:
        url (string): URL to compare with urlDatabaseList

    Returns:
       Reach based on urlDatabaseList if found otherwise it returns 1000
    """
    url = url.replace("https://", "").replace("http://", "").replace("www.", "")
    url = url[: url.find("/")]
    if url in url_list:
        return url_list[url]
    else:
        return 1000


def online_fix(df):
    """Fix the

    Args:
        df (pandas.DataFrame): DataFrame from onlineandsocial.xlsx

    Returns:
        df (pandas.DataFrame): Fix DataFrame
    """

    start_time = time.time()

    url_reach_list = url_reach()

    deleteList = populate_delete_list()

    df = df.rename(
        columns={
            "favorite": "category",
            "tags": "AVE",
        }
    )

    #  Capatilize first letter of Sentiment
    df.tone = df.tone.str.capitalize()

    #  Converts the values to datetime values
    df.published_at = pd.to_datetime(df.published_at)
    #  Gets rid of the time without time
    df.published_at = df.published_at.dt.date

    dropList = []

    for index, row in df.iterrows():
        df.loc[[index], "source_url"] = (
            row.source_url.replace("http://", "")
            .replace("https://", "")
            .replace("www.", "")
        )

        if check_url(row.source_url, deleteList):
            dropList.append(index)

    df.drop(dropList, inplace=True)

    for index, row in df.iterrows():
        if find_category(row.title, row.description, row.url):
            df.loc[[index], "category"] = "Product"
        else:
            df.loc[[index], "category"] = "Corporate"

        if (row.domain_reach == 0) or (np.isnan(row.domain_reach)):
            df.loc[[index], "domain_reach"] = find_domain(row.url, url_reach_list)

    df.AVE = df.domain_reach * 0.21

    df = df.drop(
        [
            "id",
            "alert_id",
            "parent_id",
            "children",
            "parent_url",
            "direct_reach",
            "cumulative_reach",
            "alert_name",
        ],
        axis=1,
    )

    print("--- Online Took %s seconds ---" % (time.time() - start_time))

    return df


def online_and_social_SSA_fix(df):
    """Fix the

    Args:
        df (pandas.DataFrame): DataFrame from onlineandsocial.xlsx

    Returns:
        df (pandas.DataFrame): Fix DataFrame
    """

    start_time = time.time()

    url_reach_list = url_reach()

    df = df.rename(
        columns={
            "tags": "AVE",
        }
    )

    #  Capatilize first letter of Sentiment
    df.tone = df.tone.str.capitalize()

    #  Converts the values to datetime values
    df.published_at = pd.to_datetime(df.published_at)
    #  Gets rid of the time without time
    df.published_at = df.published_at.dt.date

    for index, row in df.iterrows():
        if ((row.domain_reach == 0) or (np.isnan(row.domain_reach))) and (
            df.source_type.iloc[0] != "twitter"
        ):
            df.loc[[index], "domain_reach"] = 1000

    if df.source_type.iloc[0] == "twitter":
        df.AVE = df.cumulative_reach * 0.11
    else:
        df.AVE = df.domain_reach * 0.21

    df = df.drop(
        [
            "id",
            "alert_id",
            "parent_id",
            "children",
            "parent_url",
            "direct_reach",
            "alert_name",
        ],
        axis=1,
    )

    print("--- SSA Took %s seconds ---" % (time.time() - start_time))

    return df


def print_fix(df):
    """Fix the

    Args:
        df (pandas.DataFrame): DataFrame from print.xlsx

    Returns:
        df (pandas.DataFrame): Fix DataFrame
    """
    start_time = time.time()

    df["Category"] = ""

    #  Determines whether the category is Product or Corporate
    category = []
    for index, row in df.iterrows():
        category = []
        subList = df.Subjects[index]
        for sub in subList.split(", "):
            try:
                if productList.index(sub) != -1:
                    cat = "Product"
                    break
            except ValueError:
                cat = "Corporate"
        category.append(cat)
        df.loc[[index], "Category"] = category[0]

    # if sub in productList:
    # df.loc[[index], 'Category'] = 'Product'
    # else:
    # df.loc[[index], 'Category'] = 'Corporate'

    #  Change Sentiment from number to String
    for index, row in df.iterrows():
        if df.Sentiment[index] == 1:
            df.loc[[index], "Sentiment"] = "Positive"
        elif df.Sentiment[index] == 0:
            df.loc[[index], "Sentiment"] = "Neutral"
        elif df.Sentiment[index] == -1:
            df.loc[[index], "Sentiment"] = "Negative"

    #  Converts the values to datetime values
    df["Referred Date"] = pd.to_datetime(df["Referred Date"])
    #  Gets rid of the time
    df["Referred Date"] = df["Referred Date"].dt.date

    df = df.drop(
        ["Scanned Date", "Reach", "Language", "Curation Date", "Client"], axis=1
    )

    print("--- Print Took %s seconds ---" % (time.time() - start_time))

    return df


def broadcast_fix_old(df):
    """Fix the

    Args:
        df (pandas.DataFrame): DataFrame from broadcast.xlsx

    Returns:
        df (pandas.DataFrame): Fix DataFrame
    """
    start_time = time.time()

    df.Date = pd.to_datetime(df.Date)
    df.Date = df.Date.dt.date

    nissan = df["Client"] == "Nissan"
    week = df["Date"] >= (datetime.date.today() - datetime.timedelta(days=60))

    df = df[nissan & week]
    df["Total AVE"] = df["Total AVE"].astype(str).str[1:]
    df["Total AVE"] = df["Total AVE"].str.replace(",", "")
    df["Total AVE"] = pd.to_numeric(df["Total AVE"])

    df = df.drop(["Tagger", "Client", "FTP Export", "Timestamp"], axis=1)
    print("--- Broadcast Took %s seconds ---" % (time.time() - start_time))

    return df


def top3Summary(excelCells, top3, ave):
    for index in range(len(top3[ave])):
        excelCells[index][0].value = top3[ave].index[index]
        excelCells[index][1].value = top3[ave, "count"][index]
        excelCells[index][2].value = top3[ave, "sum"][index]


def pivotTable(df, sources, ave):
    """ """
    pivot = pd.pivot_table(
        df, index=[sources], values=[ave], aggfunc={ave: [np.sum, "count"]}
    ).sort_values(by=(ave, "sum"), ascending=False)
    return pivot


def placeImage(wb, ws):
    img = drawing.image.Image("nissan.png")
    ws.add_image(img)

    return wb


total_time = time.time()

printFile = "data-files/print.xlsx"
broadcastFile = "data-files/broadcast.xlsx"
onlineFile = "data-files/onlineandsocial.xlsx"
onlineSSAFile = "data-files/onlineSSA.xlsx"
socialSSAFile = "data-files/socialSSA.xlsx"

outputfilename = get_outputfile_name()
outputfilenameAVE = "AVE Data.xlsx"
outputfilenameSSA = get_outputfile_nameSSA()

dfPrint = inputfilePrint(printFile)
dfOnline = inputfile(onlineFile)
dfBroadcast = inputfile(broadcastFile)
dfOnlineSSA = inputfile(onlineSSAFile)
dfSocialSSA = inputfile(socialSSAFile)

dfPrint = print_fix(dfPrint)
dfOnline = online_fix(dfOnline)
dfBroadcast = broadcast_fix_old(dfBroadcast)
dfOnlineSSA = online_and_social_SSA_fix(dfOnlineSSA)
dfSocialSSA = online_and_social_SSA_fix(dfSocialSSA)

wb = load_workbook("data-files/AVE Template.xlsx")
ws = wb["Summary"]

ws["B9"].value = (
    (datetime.date.today() - datetime.timedelta(days=7)).strftime("%d.%m.%Y")
    + " - "
    + (datetime.date.today() - datetime.timedelta(days=1)).strftime("%d.%m.%Y")
)

ws["D12"].value = dfPrint.Ave.count()
ws["E12"].value = dfPrint.Ave.sum()

ws["D13"].value = dfOnline.AVE.count()
ws["E13"].value = dfOnline.AVE.sum()

radio = dfBroadcast["Platform"] == "Radio"
tv = dfBroadcast["Platform"] == "TV"

dfRadio = dfBroadcast[radio]
dfTV = dfBroadcast[tv]

pivotPrint = pivotTable(dfPrint, "Media", "Ave")
pivotOnline = pivotTable(dfOnline, "source_url", "AVE")

if not dfTV.empty:
    ws["D14"].value = dfTV["Total AVE"].count()
    ws["E14"].value = dfTV["Total AVE"].sum()
    pivotTV = pivotTable(dfTV, "Station", "Total AVE")
    top3TV = pivotTV.head(3)
    excelCellsTV = ws["K12":"M14"]
    top3Summary(excelCellsTV, top3TV, "Total AVE")

if not dfRadio.empty:
    ws["D15"].value = dfRadio["Total AVE"].count()
    ws["E15"].value = dfRadio["Total AVE"].sum()
    pivotRadio = pivotTable(dfRadio, "Station", "Total AVE")
    top3Radio = pivotRadio.head(3)
    excelCellsRadio = ws["K18":"M20"]
    top3Summary(excelCellsRadio, top3Radio, "Total AVE")


top3Print = pivotPrint.head(3)
top3Online = pivotOnline.head(3)

excelCellsPrint = ws["G12":"I14"]
excelCellsOnline = ws["G18":"I20"]

top3Summary(excelCellsPrint, top3Print, "Ave")
top3Summary(excelCellsOnline, top3Online, "AVE")

wb = placeImage(wb, ws)
wb.save(outputfilename)

outputfile(dfPrint, dfOnline, dfBroadcast, wb, outputfilename)
outputfile(dfPrint, dfOnline, dfBroadcast, wb, outputfilenameAVE)
outputfileSSA(dfOnlineSSA, dfSocialSSA, outputfilenameSSA)

print("--- Took %s seconds to generate report ---" % (time.time() - total_time))


# TODO: convert AVE column to currency format
